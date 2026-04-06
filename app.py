import streamlit as st
import re
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import barcode
from barcode.writer import ImageWriter


st.set_page_config(
    page_title="KEHE Shorts Reorder Tool",
    layout="centered"
)

st.title("KEHE Shorts → UNIFI Reorder Tool")


def normalize_upc(raw_upc: str) -> str:
    """
    Option B UPC normalization.
    Adjust ONLY here if scan testing requires change.
    """
    digits = "".join(filter(str.isdigit, raw_upc))
    digits = digits.lstrip("0")
    return digits[:11]


uploaded_files = st.file_uploader(
    "Drag & drop KeHE invoice PDFs",
    type="pdf",
    accept_multiple_files=True
)

if st.button("Process Invoices") and uploaded_files:

    reorder_rows = []
    sellthrough_rows = []

    for file in uploaded_files:
        reader = PdfReader(file)
        text = "\n".join(page.extract_text() or "" for page in reader.pages)

        # OUT OF STOCK ITEMS (REORDERABLE)
        oos_blocks = re.split(r"OUT - OF - STOCK ITEMS", text, flags=re.I)
        for block in oos_blocks[1:]:
            for line in block.splitlines():
                m = re.search(
                    r"\s(\d{9,14})\s+([A-Z0-9 &\-]+?)\s+"
                    r"(MANUFACTURER OUT|TEMP OUT|NEW ITEM)[^\n]*",
                    line,
                    re.I,
                )
                if m:
                    reorder_rows.append([
                        normalize_upc(m.group(1)),
                        m.group(2).strip(),
                        1,
                        m.group(3).strip(),
                        file.name
                    ])

        # ITEMS NOT SHIPPED (SELL THROUGH ONLY)
        ins_blocks = re.split(r"ITEMS NOT SHIPPED", text, flags=re.I)
        for block in ins_blocks[1:]:
            for line in block.splitlines():
                m = re.search(
                    r"\s(\d{9,14})\s+([A-Z0-9 &\-]+?)\s+"
                    r"(DISCONTINUED ITEM|NOT AUTHORIZED|LIMITED SUPPLY)[^\n]*",
                    line,
                    re.I,
                )
                if m:
                    sellthrough_rows.append([
                        normalize_upc(m.group(1)),
                        m.group(2).strip(),
                        m.group(3).strip(),
                        file.name
                    ])

    wb = Workbook()
    ws = wb.active
    ws.title = "KEHE Shorts Reorder List"
    ws.append([
        "UNIFI UPC",
        "Barcode",
        "Item Description",
        "Shorted Cases",
        "Short Reason",
        "Source Invoice"
    ])

    for upc, desc, cases, reason, src in reorder_rows:
        code = barcode.get("code128", upc, writer=ImageWriter())
        path = f"{upc}.png"
        code.save(upc)

        ws.append([upc, "", desc, cases, reason, src])
        ws.add_image(XLImage(path), f"B{ws.max_row}")

    ws2 = wb.create_sheet("Sell Through / Flex (No Reorder)")
    ws2.append([
        "UPC",
        "Item Description",
        "Reason",
        "Disposition",
        "Source Invoice"
    ])

    for upc, desc, reason, src in sellthrough_rows:
        ws2.append([
            upc,
            desc,
            reason,
            "Sell through product, then flex until schematic updates",
            src
        ])

    output = "KEHE_Shorts_Reorder.xlsx"
    wb.save(output)

    with open(output, "rb") as f:
        st.download_button(
            "Download Excel",
            f,
            file_name=output
        )
